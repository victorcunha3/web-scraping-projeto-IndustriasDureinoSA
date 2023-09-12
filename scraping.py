import os
import re
import datetime
from selenium import webdriver
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from typing import List
from selenium.webdriver.common.desired_capabilities import DesiredCapabilities
from openpyxl import load_workbook, Workbook

def main():
    chromedriver = r"c:\Users\Aprendiz\Downloads\chromedriver-win64\chromedriver.exe"
    service = Service(chromedriver)
    driver = webdriver.Chrome(service=service)

    driver.get('https://xpert.com.br/atg/')
    iframes = driver.find_elements(By.TAG_NAME, 'iframe')

    if len(iframes) > 0:
        driver.switch_to.frame(iframes[0])

    wait = WebDriverWait(driver, 20)
    username_field = wait.until(EC.presence_of_element_located((By.ID, 'inputUser')))
    password_field = wait.until(EC.presence_of_element_located((By.ID, 'inputPass')))

    username_field.send_keys('')
    password_field.send_keys('')

    form = driver.find_element(By.TAG_NAME, 'form')
    form.submit()

    tanques = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.content.border-solid')))

    dados_tanques: List = []

    for tanque in tanques:
        capacidade_element = tanque.find_element(By.XPATH, './/h4[contains(text(), "Capacidade")]/b')
        litros_element = tanque.find_element(By.CSS_SELECTOR, 'h1.leitura-tanque > b')
        porcentagem_element = tanque.find_element(By.CSS_SELECTOR, 'span.ng-binding')
        h3_element = tanque.find_element(By.XPATH, './/h3')
        h3_text = h3_element.text.strip()

        capacidade_text = capacidade_element.text
        litros_text = litros_element.text
        #litros = (re.sub(r'[^\d.]', '', litros_text)) if litros_text else 0.0
        porcentagem_text = porcentagem_element.text

        data_execucao = datetime.datetime.now().strftime('%d-%m-%Y')
        hora_execucao = datetime.datetime.now().strftime('%H:%M:%S')

        #capacidade = float(re.sub(r'[^\d.]', '', capacidade_text)) if capacidade_text else 0.0

        dados = {
            "Data Execução": data_execucao,
            "Litros": litros_text,
            "Capacidade": capacidade_text,
            "Porcentagem": porcentagem_text,
            "Nome": h3_text,
            "Hora Execução": hora_execucao
        }
        dados_tanques.append(dados)

    driver.quit()

    #file_path = r"t:\FABRICA\PRODUCAO\Compartilhado\4 - SETORES\4.1 - Ensacado\hexanoTeste2.xlsx"
    file_path = r't:\FABRICA\PRODUCAO\Compartilhado\4 - SETORES\4.1 - Ensacado\hexanoDadosATUALIZADO.xlsx'
    try:
        wb = load_workbook(file_path)
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Data Execução", "Litros", "Capacidade", "Porcentagem", "Nome", "Hora Execução"])

    for dados in dados_tanques:
        sheet.append([dados["Data Execução"], dados["Litros"], dados["Capacidade"],
                      dados["Porcentagem"], dados["Nome"], dados["Hora Execução"]]
                      
                      )

    wb.save(file_path)
    print("Os dados foram exportados para o Excel com sucesso.")

if __name__ == "__main__":
    main()
