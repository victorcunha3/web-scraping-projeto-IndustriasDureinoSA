import typing
import time
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
import pprint
from openpyxl import load_workbook, Workbook

chrome_driver_path = r"C:\Users\victo\Downloads\chromedriver-win64\chromedriver-win64\chromedriver.exe"

def iniciar_navegador(url: str):
    global driver
    iniciar_chromedriver = Service(chrome_driver_path)
    driver = webdriver.Chrome(service=iniciar_chromedriver)
    driver.get(url)
    iframes = driver.find_elements(By.TAG_NAME, 'iframe')
    if iframes:
        driver.switch_to.frame(iframes[0])

def buscar_campos(usuario_id: str, senha_id: str, senha_site: str, usuario_site: str):
    wait = WebDriverWait(driver, 20)
    campo_usuario = wait.until(EC.presence_of_element_located((By.ID, usuario_id)))
    campo_senha = wait.until(EC.presence_of_element_located((By.ID, senha_id)))
    campo_usuario.send_keys(usuario_site)
    campo_senha.send_keys(senha_site)
    time.sleep(2)
    formulario = driver.find_element(By.TAG_NAME, 'form')
    formulario.submit()

def busca_tanques():
    wait = WebDriverWait(driver, 10)
    try:
        tanques = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.content.border-solid')))
    except Exception as e:
        print(f"Erro ao buscar tanques: {e}")

    time.sleep(6)
    dados_tanques: typing.List = []

    for tanque in tanques:
        try:
            capacidade_element = tanque.find_element(By.XPATH, './/h4[contains(text(), "Capacidade")]/b')
            litros_element = tanque.find_element(By.CSS_SELECTOR, 'h1.leitura-tanque > b')
            porcentagem_element = tanque.find_element(By.CSS_SELECTOR, 'span.ng-binding')
            h3_element = tanque.find_element(By.XPATH, './/h3')
            h3_text = h3_element.text.strip()

            capacidade_text = capacidade_element.text
            litros_text = litros_element.text
            porcentagem_text = porcentagem_element.text

            data_execucao = datetime.datetime.now().strftime('%d-%m-%Y')
            hora_execucao = datetime.datetime.now().strftime('%H:%M:%S')

            dados: typing.Dict = {
                "Data Execução": data_execucao,
                "Litros": litros_text,
                "Capacidade": capacidade_text,
                "Porcentagem": porcentagem_text,
                "Nome Tanque": h3_text,
                "Hora Execução": hora_execucao
            }
            dados_tanques.append(dados)

        except:
            print("Impossível encontrar tanques")

        try:
            link_relatorio_diario = wait.until(EC.presence_of_element_located((By.XPATH, '//a[contains(@href, "#/analitico") and contains(text(), "Rel. Diário")]')))
        except:
            print("IMPOSSIVEL")
        
    try:
        time.sleep(7)
        link_relatorio_diario.click()
        time.sleep(3)
        print("CLICADO")
    except Exception as e:
        print(f"Erro ao clicar no link: {e}")
    
    dados_relatorio_dia: typing.List[typing.Dict[str, str]] = []
    tanques_relatorio = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.content.border-solid')))
    for div in tanques_relatorio:
        linhas = div.text.split('\n')
        relatorio_dicionario: typing.Dict[str, str] = {}

        for linha in linhas:
            if ':' in linha:
                chave, valor = map(str.strip, linha.split(':', 1))
                relatorio_dicionario[chave] = valor
        
        dados_relatorio_dia.append(relatorio_dicionario)
    #pprint.pprint(dados_relatorio_dia)
    contador_tanque: int = 0
    relatorio_lista_final: typing.List = []

    for tanque_atual in dados_relatorio_dia:
        try:
            contador_tanque += 1
            volume_inicial_tanque = tanque_atual['Volume Inicial']
            #Não foi necessário usar a capacidade
            #capacidade_total = tanque_atual['Capacidade']
            volume_combustivel = tanque_atual["Volume Combustível"]

            dados_relatorio_final = {
                "Tanque": contador_tanque,
                "Volume inicial": volume_inicial_tanque,
                #"Capacidade": capacidade_total,
                "Volume de Combustível": volume_combustivel
            }
            relatorio_lista_final.append(dados_relatorio_final)

        except:
            print("Erro em Iteração!!")
    
    driver.quit()
    pprint.pprint(relatorio_lista_final)
    time.sleep(3)

    #Adicionando dados ao excel
    enviar_dados_excel(dados_tanques)
    enviar_relatorio_excel(relatorio_lista_final)

def enviar_dados_excel(dados_tanques):
    try:
        wb = load_workbook('tanque_excel.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Data Execução", "Litros", "Capacidade", "Porcentagem", "Nome", "Hora Execução"])

    for dados in dados_tanques:
        sheet.append([dados["Data Execução"], dados["Litros"], dados["Capacidade"],
                      dados["Porcentagem"], dados["Nome Tanque"], dados["Hora Execução"]]
                      
                      )

    wb.save('tanque_excel.xlsx')
    print("Os dados foram exportados para o Excel com sucesso.")


def enviar_relatorio_excel(relatorio):
    try:
        wb = load_workbook('relatorio_excel.xlsx')
        sheet = wb.active
    except FileNotFoundError:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Tanque", "Volume Inicial", "Volume de Combustível"])
    
    for dado_atual in relatorio:
        sheet.append([dado_atual["Tanque"], dado_atual["Volume inicial"],
                      dado_atual["Volume de Combustível"]])
    
    wb.save('relatorio_excel.xlsx')
    print("Os dados de relatório exportados para o Excel com sucesso.")


def main():
    iniciar_navegador("https://xpert.com.br/atg/")
    buscar_campos("inputUser", "inputPass", "", "")
    busca_tanques()
    driver.quit()


if __name__ == "__main__":
    main()
