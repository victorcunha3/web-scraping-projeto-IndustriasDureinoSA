import typing
import time
import datetime
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from openpyxl import load_workbook, Workbook
import pprint
import re

# Const
CHROME_DRIVER_PATH:str = r"c:\Users\aprendiz.DUREINO\Downloads\driverCurrent\chromedriver-win64\chromedriver.exe"
TANQUE_EXCEL_FILE:str = 'tanque_excel.xlsx'
RELATORIO_EXCEL_FILE:str = 'estoque_hexano.xlsx'

class TanqueHexano:
    #Realizar encapsulamento OPCIONAL
    def __init__(self, url: str, usuario_id: str, senha_id: str, senha_site: str, usuario_site: str):
        self.url = url
        self.usuario_id = usuario_id
        self.senha_id = senha_id
        self.senha_site = senha_site
        self.usuario_site = usuario_site
        self.driver = None

    def iniciar_navegador(self) -> None:
        iniciar_chromedriver = Service(CHROME_DRIVER_PATH)
        self.driver = webdriver.Chrome(service=iniciar_chromedriver)
        self.driver.get(self.url)
        iframes = self.driver.find_elements(By.TAG_NAME, 'iframe')
        if iframes:
            self.driver.switch_to.frame(iframes[0])

    def buscar_campos(self) -> None:
        wait = WebDriverWait(self.driver, 20)
        campo_usuario = wait.until(EC.presence_of_element_located((By.ID, self.usuario_id)))
        campo_senha = wait.until(EC.presence_of_element_located((By.ID, self.senha_id)))
        campo_usuario.send_keys(self.usuario_site)
        campo_senha.send_keys(self.senha_site)
        time.sleep(2)
        formulario = self.driver.find_element(By.TAG_NAME, 'form')
        formulario.submit()

    def busca_tanques(self) -> None:
        wait = WebDriverWait(self.driver, 10)
        try:
            tanques = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.content.border-solid')))
        except Exception as e:
            print(f"Erro ao buscar tanques: {e}")

        time.sleep(6)
        dados_tanques: typing.List = []

        for tanque_atual in tanques:
            try:
                capacidade_element = tanque_atual.find_element(By.XPATH, './/h4[contains(text(), "Capacidade")]/b')
                litros_element = tanque_atual.find_element(By.CSS_SELECTOR, 'h1.leitura-tanque > b')
                porcentagem_element = tanque_atual.find_element(By.CSS_SELECTOR, 'span.ng-binding')
                h3_element = tanque_atual.find_element(By.XPATH, './/h3')
                h3_text = h3_element.text.strip()

                capacidade_text = capacidade_element.text
                litros_text = litros_element.text
                porcentagem_text = porcentagem_element.text

                data_execucao = datetime.datetime.now().strftime('%d-%m-%Y')
                hora_execucao = datetime.datetime.now().strftime('%H:%M:%S')

                dados: typing.Dict = {
                    "Data Execução": data_execucao,
                    "Litros": litros_text,
                    "Capacidade": capacidade_text,
                    "Porcentagem": porcentagem_text,
                    "Nome Tanque": h3_text,
                    "Hora Execução": hora_execucao
                }
                dados_tanques.append(dados)

            except Exception as e:
                print(f"Impossível encontrar tanques: {e}")

        try:
            time.sleep(7)
            link_relatorio_diario = wait.until(EC.presence_of_element_located(
                (By.XPATH, '//a[contains(@href, "#/analitico") and contains(text(), "Rel. Diário")]')))
            #print(link_relatorio_diario)
        except Exception as e:
            print(f"IMPOSSIVEL encontrar link relatório diário: {e}")


        try:
            time.sleep(7)
            link_relatorio_diario.click()
            time.sleep(3)
            print("CLICADO")
        except Exception as e:
            print(f"Erro ao clicar no link: {e}")

        dados_relatorio_dia: typing.List[typing.Dict[str, str]] = []
        tanques_relatorio = wait.until(EC.presence_of_all_elements_located((By.CSS_SELECTOR, 'div.content.border-solid')))

        chave = None
        for div in tanques_relatorio:
            linhas = div.text.split('\n')
            #pprint.pprint(linhas)
            relatorio_dicionario = {}

            for dado_atual in linhas:
                if dado_atual.endswith(":"):
                    chave = dado_atual.strip(":").strip()

                elif chave is not None:
                    valor = dado_atual.strip()
                    relatorio_dicionario[chave] = valor
                    chave = None
                    valor = None
            dados_relatorio_dia.append(relatorio_dicionario)

        contador_tanque: int = 0
        relatorio_lista_final: typing.List = []
        pprint.pprint(dados_relatorio_dia)

        for tanque_atual in dados_relatorio_dia:
            data_execucao = datetime.datetime.now().strftime('%d-%m-%Y')
            hora_execucao = datetime.datetime.now().strftime('%H:%M:%S')
            try:
                contador_tanque += 1
                volume_inicial_tanque = tanque_atual['Volume Inicial']
                #volume_combustivel = tanque_atual["Volume Combustível"]

                dados_relatorio_final = {
                    "Tanque": contador_tanque,
                    "Volume inicial": volume_inicial_tanque,
                    #"Volume de Combustível": volume_combustivel,
                    "Data Execução": data_execucao,
                    "Hora Execução": hora_execucao
                }
                relatorio_lista_final.append(dados_relatorio_final)

            except Exception as e:
                print(f"Erro em Iteração!! {e}")

        self.driver.quit()
        #pprint.pprint(relatorio_lista_final)
        time.sleep(3)

        # Adicionando dados ao excel
        self.enviar_dados_excel(dados_tanques)
        self.enviar_relatorio_excel(relatorio_lista_final)

    def enviar_dados_excel(self, dados_tanques: typing.List) -> None:
        try:
            wb = load_workbook(TANQUE_EXCEL_FILE)
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append(["Data Execução", "Litros", "Capacidade", "Porcentagem", "Nome", "Hora Execução"])

        for dados in dados_tanques:
            sheet.append([dados["Data Execução"], dados["Litros"], dados["Capacidade"],
                          dados["Porcentagem"], dados["Nome Tanque"], dados["Hora Execução"]])

        wb.save(TANQUE_EXCEL_FILE)
        print("Os dados foram exportados para o Excel com sucesso.")

    def enviar_relatorio_excel(self, relatorio: typing.List) -> None:
        try:
            wb = load_workbook(RELATORIO_EXCEL_FILE)
            sheet = wb.active
        except FileNotFoundError:
            wb = Workbook()
            sheet = wb.active
            sheet.append(["Tanque", "Volume Inicial", "Data execução", "Hora execução"])

        for dado_atual in relatorio:
            sheet.append([dado_atual["Tanque"], dado_atual["Volume inicial"],
                        dado_atual["Data Execução"],
                        dado_atual["Hora Execução"]])

        wb.save(RELATORIO_EXCEL_FILE)
        print("Os dados de relatório foram exportados para o Excel com sucesso.")

def main():
    tanques_hexano_class = TanqueHexano("https://xpert.com.br/atg/",
                                  "inputUser",
                                  "inputPass",
                                  "",
                                  "")

    tanques_hexano_class.iniciar_navegador()
    tanques_hexano_class.buscar_campos()
    tanques_hexano_class.busca_tanques()

if __name__ == "__main__":
    main()
